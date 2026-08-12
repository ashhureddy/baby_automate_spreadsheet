"""
Streamlit app for Advanced Cellular Template Processing - LOCAL OCR EDITION
Place this file as app.py in your repository. No API keys required.
"""

import os
import io
import re
import json
import time
import tempfile
from pathlib import Path
from typing import Optional, List

import streamlit as st
import openpyxl
from PIL import Image
import cv2
import pytesseract

# ---------------- Schemas ----------------
SERVICE_SCHEMA = {
    "nr_arfcn": "number", "nr_band": "number", "nr_pci": "number", "nr_bw": "number",
    "nr5g_rsrp": "number", "nr5g_rsrq": "number", "nr5g_sinr": "number",
    "lte_band": "number", "lte_earfcn": "number", "lte_pci": "number", "lte_bw": "number",
    "lte_rsrp": "number", "lte_rsrq": "number", "lte_sinr": "number",
}

# ---------------- Globals ----------------
alpha_service, beta_service, gamma_service = {}, {}, {}
alpha_speedtest, beta_speedtest, gamma_speedtest = {}, {}, {}
alpha_video, beta_video, gamma_video = {}, {}, {}
voice_test, extract_text, avearge = {}, [], {}

def log_append(log_placeholder, logs_list: list, msg: str):
    ts = time.strftime("%H:%M:%S")
    line = f"[{ts}] {msg}"
    logs_list.append(line)
    display = "\n".join(logs_list[-2000:])
    try:
        log_placeholder.text_area("Logs", value=display, height=360)
    except Exception:
        pass

def get_sector_from_col(col_index: int) -> str:
    if 0 <= col_index < 4: return "alpha"
    if 4 <= col_index < 8: return "beta"
    if 8 <= col_index < 12: return "gamma"
    if 12 <= col_index < 18: return "voicetest"
    return "unknown"

# ---------------- OCR Engine Helpers ----------------
def extract_text_from_image(img_path: str, psm: int = 3, digit_only: bool = False) -> str:
    img = cv2.imread(img_path)
    if img is None: return ""
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    gray = cv2.resize(gray, None, fx=2, fy=2, interpolation=cv2.INTER_CUBIC)
    
    config = f'--psm {psm}'
    if digit_only:
        config += ' -c tessedit_char_whitelist=0123456789.-'
        
    try:
        return pytesseract.image_to_string(gray, config=config)
    except pytesseract.TesseractNotFoundError:
        return ""

def extract_text_from_crop(img_path: str, crop_box: tuple, digit_only: bool = True) -> str:
    img = cv2.imread(img_path)
    if img is None: return ""
    
    h, w = img.shape[:2]
    x1, y1 = int(w * crop_box[0]), int(h * crop_box[1])
    x2, y2 = int(w * crop_box[2]), int(h * crop_box[3])
    
    roi = img[y1:y2, x1:x2]
    gray = cv2.cvtColor(roi, cv2.COLOR_BGR2GRAY)
    gray = cv2.resize(gray, None, fx=2.5, fy=2.5, interpolation=cv2.INTER_CUBIC)
    
    config = '--psm 7'
    if digit_only: config += ' -c tessedit_char_whitelist=0123456789.-'
        
    try:
        return pytesseract.image_to_string(gray, config=config).strip()
    except pytesseract.TesseractNotFoundError:
        return ""

# ---------------- Analysis Modules ----------------
def process_service_images_local(image_paths: list, log_placeholder, logs: list) -> dict:
    data = {}
    full_text = ""
    for path in image_paths:
        if path: full_text += extract_text_from_image(path) + "\n"
        
    # LTE Regex (Jumps over noisy colons/spaces)
    m = re.search(r'Earfcn[^\d]*(\d+)', full_text, re.IGNORECASE)
    if m: data['lte_earfcn'] = int(m.group(1))
    m = re.search(r'PCI[^\d]*(\d+)', full_text, re.IGNORECASE)
    if m: data['lte_pci'] = int(m.group(1))
    m = re.search(r'LTE.*?BAND[^\d]*(\d+)', full_text, re.IGNORECASE)
    if m: data['lte_band'] = int(m.group(1))
    m = re.search(r'LTE.*?BW[^\d]*(\d+)', full_text, re.IGNORECASE)
    if m: data['lte_bw'] = int(m.group(1))
    m = re.search(r'RSRP[^\d\-]*(-?\d+)', full_text, re.IGNORECASE)
    if m: data['lte_rsrp'] = int(m.group(1))
    m = re.search(r'RSRQ[^\d\-]*(-?\d+)', full_text, re.IGNORECASE)
    if m: data['lte_rsrq'] = int(m.group(1))
    m = re.search(r'SNR[^\d\-]*([\d\.-]+)', full_text, re.IGNORECASE)
    if m: data['lte_sinr'] = float(m.group(1))
        
    # NR Regex
    m = re.search(r'NR5G_RSRP[^\d\-]*(-?\d+)', full_text, re.IGNORECASE)
    if m: data['nr5g_rsrp'] = int(m.group(1))
    m = re.search(r'NR5G_SINR[^\d\-]*([\d\.-]+)', full_text, re.IGNORECASE)
    if m: data['nr5g_sinr'] = float(m.group(1))
    m = re.search(r'NR5G RSRQ[^\d\-]*(-?\d+)', full_text, re.IGNORECASE)
    if m: data['nr5g_rsrq'] = int(m.group(1))
    m = re.search(r'NR_ARFCN[^\d]*(\d+)', full_text, re.IGNORECASE)
    if m: data['nr_arfcn'] = int(m.group(1))
    m = re.search(r'NR_PCI[^\d]*(\d+)', full_text, re.IGNORECASE)
    if m: data['nr_pci'] = int(m.group(1))
    m = re.search(r'NR_BAND[^\d]*[nN]?(\d+)', full_text, re.IGNORECASE)
    if m: data['nr_band'] = int(m.group(1))
    m = re.search(r'NR_BW[^\d]*(\d+)', full_text, re.IGNORECASE)
    if m: data['nr_bw'] = int(m.group(1))

    return data

def analyze_speed_test_local(image_path: str, log_placeholder, logs: list) -> Optional[dict]:
    # 1. Primary Method: Extract all text and use robust Regex
    full_text = extract_text_from_image(image_path, psm=3)
    # Strip commas so numbers like 1,024 are read cleanly as 1024
    clean_text = full_text.replace(',', '')
    
    dl, ul, ping = None, None, None
    
    # Look for "Download" (and optional "Mbps"), then grab the next sequence of digits
    dl_match = re.search(r'Download.*?Mbps[\s\n]*([\d\.]+)', clean_text, re.IGNORECASE)
    if not dl_match:
        dl_match = re.search(r'Download[\s\n]*([\d\.]+)', clean_text, re.IGNORECASE)
    if dl_match: dl = dl_match.group(1)

    # Look for "Upload"
    ul_match = re.search(r'Upload.*?Mbps[\s\n]*([\d\.]+)', clean_text, re.IGNORECASE)
    if not ul_match:
        ul_match = re.search(r'Upload[\s\n]*([\d\.]+)', clean_text, re.IGNORECASE)
    if ul_match: ul = ul_match.group(1)
    
    # Look for "Ping"
    ping_match = re.search(r'Ping.*?ms[\s\n]*(\d+)', clean_text, re.IGNORECASE)
    if ping_match: ping = ping_match.group(1)

    # 2. Fallback Method: Widened Bounding Boxes if Regex missed it
    if not dl or not ul:
        dl_text = extract_text_from_crop(image_path, (0.02, 0.30, 0.48, 0.60)) 
        ul_text = extract_text_from_crop(image_path, (0.50, 0.30, 0.98, 0.60))
        ping_text = extract_text_from_crop(image_path, (0.05, 0.60, 0.40, 0.75))
        
        # Clean any non-numeric noise the boxes might have picked up
        if not dl and dl_text: dl = re.sub(r'[^\d\.]', '', dl_text)
        if not ul and ul_text: ul = re.sub(r'[^\d\.]', '', ul_text)
        if not ping and ping_text: ping = re.sub(r'[^\d\.]', '', ping_text)

    try:
        dl_val = float(dl) if dl else None
        ul_val = float(ul) if ul else None
        
        # Prevent video screenshots from being logged as speed tests
        if dl_val in [2160, 1080, 720]: return None 
        
        return {
            "image_type": "speed_test",
            "data": {
                "download_mbps": dl_val,
                "upload_mbps": ul_val,
                "ping_ms": int(float(ping)) if ping else None,
            }
        }
    except Exception:
        return None

def analyze_video_test_local(image_path: str, log_placeholder, logs: list) -> Optional[dict]:
    full_text = extract_text_from_image(image_path)
    
    # Catch 2160, 1080, 720, or 4K regardless of formatting
    res_match = re.search(r'(2160|1080|720|1440|4K)', full_text, re.IGNORECASE)
    load_match = re.search(r'Load.*?Time[^\d]*(\d+)', full_text, re.IGNORECASE)
    buf_match = re.search(r'Buffering[^\d]*(\d+)', full_text, re.IGNORECASE)
    
    if not res_match and not load_match: return None
    
    resolution = res_match.group(1) if res_match else None
    if resolution == "4K": resolution = "2160"
    
    return {
        "image_type": "video_test",
        "data": {
            "max_resolution": f"{resolution}p" if resolution else None,
            "load_time_ms": int(load_match.group(1)) if load_match else None,
            "buffering_percentage": int(buf_match.group(1)) if buf_match else 0
        }
    }

def analyze_voice_test_local(image_path: str, log_placeholder, logs: list) -> Optional[dict]:
    full_text = extract_text_from_image(image_path)
    
    # Find all timestamps (e.g. 12:48 or 01:05) in the text
    time_matches = re.findall(r'(\d{1,2})[;:\.](\d{2})', full_text)
    
    extracted_time = None
    duration = None
    
    if len(time_matches) >= 2:
        # First match is usually the phone clock (Time), second is the Call Duration
        extracted_time = f"{time_matches[0][0]}:{time_matches[0][1]}"
        duration = int(time_matches[1][0]) * 60 + int(time_matches[1][1])
    elif len(time_matches) == 1:
        # If only one is found, assume it's the duration
        duration = int(time_matches[0][0]) * 60 + int(time_matches[0][1])

    # Fallback for Duration using cropping if text extraction missed it entirely
    if duration is None:
        dur_text = extract_text_from_crop(image_path, (0.35, 0.05, 0.65, 0.15), digit_only=False)
        m = re.search(r'(\d{1,2})[;:\.](\d{2})', dur_text)
        if m:
            duration = int(m.group(1)) * 60 + int(m.group(2))
            
    return {
        "image_type": "voice_call",
        "data": {
            "time": extracted_time,
            "call_duration_seconds": duration, 
            "call_status": "active" if duration else "dialing"
        }
    }

def dispatch_image_analysis_local(image_path: str, log_placeholder, logs: list) -> Optional[dict]:
    try: idx = int(Path(image_path).stem.split("_")[-1])
    except: idx = 0

    if 3 <= idx <= 7:
        priority = [analyze_speed_test_local, analyze_video_test_local, analyze_voice_test_local]
    elif idx >= 8:
        priority = [analyze_video_test_local, analyze_speed_test_local, analyze_voice_test_local]
    else:
        priority = [analyze_speed_test_local, analyze_video_test_local, analyze_voice_test_local]

    for func in priority:
        res = func(image_path, log_placeholder, logs)
        if res and res.get("data"):
            if res["image_type"] == "speed_test" and res["data"].get("download_mbps") is None: continue
            return res
    return None

# ---------------- Excel Processing logic ----------------
def extract_images_from_excel(xlsx_path: str, output_folder: str, log_placeholder, logs: list) -> List[str]:
    wb = openpyxl.load_workbook(xlsx_path)
    sheet = wb.active
    images = getattr(sheet, "_images", [])
    os.makedirs(output_folder, exist_ok=True)
    images_with_locations = []
    
    for image in images:
        try: row, col = image.anchor._from.row + 1, image.anchor._from.col
        except: row, col = 0, 0
        images_with_locations.append({"image": image, "row": row, "col": col})

    images_sorted = sorted(images_with_locations, key=lambda i: (i["row"], i["col"]))
    saved_paths = []
    counters = {"alpha": 0, "beta": 0, "gamma": 0, "voicetest": 0, "unknown": 0}

    for itm in images_sorted:
        sector = get_sector_from_col(itm["col"])
        counters[sector] += 1
        filename = f"{sector}_image_{counters[sector]}.png"
        out_path = os.path.join(output_folder, filename)
        img_data = itm["image"]._data()
        pil = Image.open(io.BytesIO(img_data)).convert("RGB")
        pil.save(out_path, "PNG")
        saved_paths.append(out_path)
    return saved_paths

def _normalize_name(s: str) -> str: return re.sub(r"[^0-9a-zA-Z]", "", s).lower()

# UPDATED: More robust regex to handle keys with or without quotes
key_pattern = re.compile(r"\[['\"]?([^'\"\]]+)['\"]?\]")

def resolve_expression_with_vars(expr: str, allowed_vars: dict):
    expr = expr.strip()
    m = re.match(r"^([A-Za-z_]\w*)(.*)$", expr)
    if not m: return None
    base_raw, rest = m.group(1), m.group(2) or ""

    norm_map = {_normalize_name(k): k for k in allowed_vars.keys()}
    base_key = norm_map.get(_normalize_name(base_raw))
    if not base_key: return None

    obj = allowed_vars[base_key]
    if rest.strip() == "": return obj

    keys = key_pattern.findall(rest)
    try:
        for k in keys:
            found = None
            for real_k in obj.keys():
                if real_k.lower() == k.lower() or _normalize_name(real_k) == _normalize_name(k):
                    found = real_k
                    break
            if found: obj = obj[found]
            else: return None
        return obj
    except: return None

def process_file_streamlit(user_file_path: str, temp_dir: str, logs: list, text_area_placeholder) -> Optional[str]:
    global alpha_service, beta_service, gamma_service, alpha_speedtest, beta_speedtest, gamma_speedtest
    global alpha_video, beta_video, gamma_video, voice_test, avearge

    images_temp = os.path.join(temp_dir, "images")
    image_paths = extract_images_from_excel(user_file_path, images_temp, text_area_placeholder, logs)

    images_by_sector = {"alpha": [], "beta": [], "gamma": [], "voicetest": [], "unknown": []}
    for p in image_paths: images_by_sector[Path(p).stem.split("_")[0]].append(p)

    log_append(text_area_placeholder, logs, "[LOG] Starting Local OCR mapping...")
    
    for sector in ["alpha", "beta", "gamma"]:
        sector_images = images_by_sector[sector]
        img1 = next((p for p in sector_images if Path(p).stem.endswith("_1")), None)
        img2 = next((p for p in sector_images if Path(p).stem.endswith("_2")), None)

        svc = process_service_images_local([img1, img2], text_area_placeholder, logs)
        if sector == "alpha": alpha_service = svc
        elif sector == "beta": beta_service = svc
        elif sector == "gamma": gamma_service = svc

        for img in [p for p in sector_images if p not in (img1, img2)]:
            res = dispatch_image_analysis_local(img, text_area_placeholder, logs)
            if res:
                name = Path(img).stem
                if res["image_type"] == "speed_test":
                    {"alpha": alpha_speedtest, "beta": beta_speedtest, "gamma": gamma_speedtest}[sector][name] = res["data"]
                elif res["image_type"] == "video_test":
                    {"alpha": alpha_video, "beta": beta_video, "gamma": gamma_video}[sector][name] = res["data"]

    for img in images_by_sector["voicetest"]:
        res = analyze_voice_test_local(img, text_area_placeholder, logs)
        if res: voice_test[Path(img).stem] = res["data"]

    wb = openpyxl.load_workbook(user_file_path)
    sheet = wb.active
    
    allowed_vars = {
        "alpha_service": alpha_service, "beta_service": beta_service, "gamma_service": gamma_service,
        "alpha_speedtest": alpha_speedtest, "beta_speedtest": beta_speedtest, "gamma_speedtest": gamma_speedtest,
        "alpha_video": alpha_video, "beta_video": beta_video, "gamma_video": gamma_video,
        "voice_test": voice_test, "avearge": avearge,
    }

    def _is_red(font):
        if not font or not getattr(font, "bold", False): return False
        col = getattr(font, "color", None)
        return str(getattr(col, "rgb", "")).upper().endswith("FF0000")

    # Safe mapping block
    for row in sheet.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and _is_red(cell.font):
                expr = cell.value.strip()
                
                # Only strip outer quotes if the whole expression was accidentally wrapped
                if (expr.startswith('"') and expr.endswith('"')) or (expr.startswith("'") and expr.endswith("'")):
                    expr = expr[1:-1].strip()
                
                resolved = resolve_expression_with_vars(expr, allowed_vars)
                
                if resolved is not None:
                    # Convert dicts/lists to strings so Excel doesn't crash
                    if isinstance(resolved, (dict, list, tuple)):
                        try:
                            cell.value = json.dumps(resolved)
                        except Exception:
                            cell.value = str(resolved)
                    else:
                        cell.value = resolved
                else:
                    cell.value = "NULL"

    wb.save(user_file_path)
    log_append(text_area_placeholder, logs, "[SUCCESS] Local OCR Complete. File saved.")
    return user_file_path

# ---------------- UI ----------------
def main_ui():
    st.set_page_config(page_title="Cellular Template Processor (Local)", layout="wide")
    st.title("Advanced Cellular Template Processor (100% Free / Local)")
    
    try:
        pytesseract.get_tesseract_version()
    except pytesseract.TesseractNotFoundError:
        st.error("🚨 Tesseract OCR is not installed on this system! Ensure you have rebuilt the Dev Container or installed 'tesseract-ocr' via apt-get.")

    if "logs" not in st.session_state: st.session_state["logs"] = []
    log_placeholder = st.empty()
    
    uploaded_file = st.file_uploader("Upload .xlsx template", type=["xlsx"])

    if uploaded_file:
        tmp_dir = tempfile.mkdtemp()
        saved_path = os.path.join(tmp_dir, uploaded_file.name)
        with open(saved_path, "wb") as f: f.write(uploaded_file.read())

        if st.button("Process file now"):
            out_path = process_file_streamlit(saved_path, tmp_dir, st.session_state["logs"], log_placeholder)
            if out_path:
                with open(out_path, "rb") as f:
                    st.download_button("Download Processed File", data=f, file_name=f"Processed_{uploaded_file.name}")

if __name__ == "__main__":
    main_ui()
